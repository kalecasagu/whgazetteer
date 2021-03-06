#from django.core import serializers
from django.http import FileResponse, JsonResponse # HttpResponse
from django.shortcuts import get_object_or_404, render
#from django.utils.encoding import smart_str
from django.views.generic import View

import codecs, sys, os, pprint, time, datetime, csv, openpyxl
import simplejson as json
from goodtables import validate as gvalidate
from jsonschema import draft7_format_checker, validate #,Draft7Validator
from pathlib import Path
from shapely import wkt

from datasets.models import Dataset, DatasetUser, Hit
from datasets.static.hashes import aat, parents, aat_q
from places.models import PlaceGeom # ,Place
pp = pprint.PrettyPrinter(indent=1)

# ***
# UPLOAD UTILS
# ***
def xl_tester(): 
  fn = '/Users/karlg/repos/_whgdata/data/_source/CentralEurasia/bregel_in progress.xlsx'
  from openpyxl import load_workbook
  wb = load_workbook(filename = fn)
  sheet_ranges = wb['range names']

def xl_upload(request):
  if "GET" == request.method:
    return render(request, 'datasets/xl.html', {})
  else:
    excel_file = request.FILES["excel_file"]

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting all sheets
    sheets = wb.sheetnames
    print(sheets)

    # getting a particular sheet by name out of many sheets
    ws = wb["Sheet1"]
    print(ws)

    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    for row in ws.iter_rows():
      row_data = list()
      for cell in row:
        #row_data.append(str(cell.value))
        row_data.append(cell.value)
      excel_data.append(row_data)

    return render(request, 'datasets/xl.html', {"excel_data":excel_data}) 

# ***
# DOWNLOAD FILES
# ***
# TODO: use DRF serializer? download_{format} methods on api.PlaceList() view?
# https://stackoverflow.com/questions/38697529/how-to-return-generated-file-download-with-django-rest-framework
def download_file(request, *args, **kwargs):
  ds=get_object_or_404(Dataset,pk=kwargs['id'])
  fileobj = ds.files.all().order_by('-rev')[0]
  fn = 'media/'+fileobj.file.name
  file_handle = fileobj.file.open()
  print('download_file: kwargs,fn,fileobj.format',kwargs,fn,fileobj.format)
  # set content type
  response = FileResponse(file_handle, content_type='text/csv' if fileobj.format=='delimited' else 'text/json')
  response['Content-Disposition'] = 'attachment; filename="'+fileobj.file.name+'"'

  return response
# TODO: crude, only properties are ids
def download_gis(request, *args, **kwargs):
  print('download_gis kwargs',kwargs)
  user = request.user.username
  ds=get_object_or_404(Dataset,pk=kwargs['id'])
  date=maketime()
  # make file name
  fn = 'media/user_'+user+'/'+ds.label+'_aug_'+date+'.json'
  # open it for write
  fout = codecs.open(fn,'w','utf8')

  # build a flat FeatureCollection 
  features=PlaceGeom.objects.filter(place_id__in=ds.placeids).values_list('jsonb','place_id','src_id')
  fcoll = {"type":"FeatureCollection","features":[]}
  for f in features:
    feat={"type":"Feature","properties":{"pid":f[1],"src_id":f[2]},"geometry":f[0]}
    fcoll['features'].append(feat)
  fout.write(json.dumps(fcoll))
  fout.close()
  # response is reopened file and content type
  response = FileResponse(open(fn, 'rb'),content_type='text/json')
  response['Content-Disposition'] = 'attachment; filename="mydata.geojson"'

  return response
  #return HttpResponse(content='download_gis: '+fn)

# returns file in original format w/any new geoms, links
def download_augmented(request, *args, **kwargs):
  from django.db import connection
  print('download_augmented kwargs',kwargs)
  user = request.user.username
  ds=get_object_or_404(Dataset,pk=kwargs['id'])
  dslabel = ds.label
  fileobj = ds.files.all().order_by('-rev')[0]
  date=maketime()

  req_format = kwargs['format']
  if req_format is not None:
    print('got format',req_format)
    #qs = qs.filter(title__icontains=query)

  features=ds.places.all() 

  if fileobj.format == 'delimited' and req_format == 'tsv':
    print('augmented for delimited')
    # make file name
    fn = 'media/user_'+user+'/'+ds.label+'_aug_'+date+'.tsv'
    def augLinks(linklist):
      aug_links = []
      for l in linklist:
        aug_links.append(l.jsonb['identifier'])
      return ';'.join(aug_links)

    def augGeom(qs_geoms):
      gobj = {'new':[]}
      for g in qs_geoms:
        if not g.task_id:
          # it's an original
          gobj['lonlat'] = g.jsonb['coordinates']
        else:
          # it's an aug/add
          gobj['new'].append({"id":g.jsonb['citation']['id'],"coordinates":g.jsonb['coordinates']})
      return gobj

    with open(fn, 'w', newline='', encoding='utf-8') as csvfile:
      writer = csv.writer(csvfile, delimiter='\t', quotechar='', quoting=csv.QUOTE_NONE)
      writer.writerow(['id','whg_pid','title','ccodes','lon','lat','added','matches'])
      for f in features:
        geoms = f.geoms.all()
        gobj = augGeom(geoms)
        row = [str(f.src_id),
               str(f.id),f.title,
               ';'.join(f.ccodes),
               gobj['lonlat'][0] if 'lonlat' in gobj else None,
               gobj['lonlat'][1] if 'lonlat' in gobj else None,
               gobj['new'] if 'new' in gobj else None,
               str(augLinks(f.links.all()))
               ]
        writer.writerow(row)
        #print(row)
    response = FileResponse(open(fn, 'rb'),content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="'+os.path.basename(fn)+'"'

    return response
  else:
    print('download_augmented()')
    # make file name
    fn = 'media/user_'+user+'/'+ds.label+'_aug_'+date+'.json'
    result={"type":"FeatureCollection","features":[]}
    with open(fn, 'w', encoding='utf-8') as outfile:
      with connection.cursor() as cursor:
        cursor.execute("""with namings as 
          (select place_id, jsonb_agg(jsonb) as names from place_name pn 
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          placetypes as 
          (select place_id, jsonb_agg(jsonb) as "types" from place_type pt 
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          placelinks as 
          (select place_id, jsonb_agg(jsonb) as links from place_link pl 
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          geometry as 
          (select place_id, jsonb_agg(jsonb) as geoms from place_geom pg 
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          placewhens as
          (select place_id, jsonb_agg(jsonb) as whens from place_when pw 
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          placerelated as
          (select place_id, jsonb_agg(jsonb) as rels from place_related pr 
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          descriptions as
          (select place_id, jsonb_agg(jsonb) as descrips from place_description pdes 
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          depictions as
          (select place_id, jsonb_agg(jsonb) as depicts from place_depiction pdep 
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id )	
          select jsonb_build_object(
            'type','Feature',
            'properties', jsonb_build_object(
                'id',p.id,'src_id',p.src_id),
            'names',n.names,
            'types',pt.types,
            'links',pl.links,
            'geometry',jsonb_build_object(
                'type','GeometryCollection',
                'geometries',g.geoms),
            'whens',pw.whens,
            'related',pr.rels,
            'descriptions',pdes.descrips,
            'depictions',pdep.depicts
          ) from places p 
          left join namings n on p.id = n.place_id
          left join placetypes pt on p.id = pt.place_id
          left join placelinks pl on p.id = pl.place_id
          left join geometry g on p.id = g.place_id
          left join placewhens pw on p.id = pw.place_id
          left join placerelated pr on p.id = pr.place_id
          left join descriptions pdes on p.id = pdes.place_id
          left join depictions pdep on p.id = pdep.place_id
          where dataset = '{ds}'        
        """.format(ds=dslabel))
        for row in cursor:
          result['features'].append(row[0])
        outfile.write(json.dumps(result,indent=2))
    # response is reopened file
    response = FileResponse(open(fn, 'rb'), content_type='text/json')
    response['Content-Disposition'] = 'attachment; filename="'+os.path.basename(fn)+'"'

    return response

def download_augmented_slow(request, *args, **kwargs):
  print('download_augmented kwargs',kwargs)
  user = request.user.username
  ds=get_object_or_404(Dataset,pk=kwargs['id'])
  fileobj = ds.files.all().order_by('-rev')[0]
  date=maketime()

  req_format = kwargs['format']
  if req_format is not None:
    print('got format',req_format)
    #qs = qs.filter(title__icontains=query)

  features=ds.places.all() 

  if fileobj.format == 'delimited' and req_format == 'tsv':
    print('augmented for delimited')
    # make file name
    fn = 'media/user_'+user+'/'+ds.label+'_aug_'+date+'.tsv'
    def augLinks(linklist):
      aug_links = []
      for l in linklist:
        aug_links.append(l.jsonb['identifier'])
      return ';'.join(aug_links)

    def augGeom(qs_geoms):
      gobj = {'new':[]}
      for g in qs_geoms:
        if not g.task_id:
          # it's an original
          gobj['lonlat'] = g.jsonb['coordinates']
        else:
          # it's an aug/add
          gobj['new'].append({"id":g.jsonb['citation']['id'],"coordinates":g.jsonb['coordinates']})
      return gobj

    with open(fn, 'w', newline='', encoding='utf-8') as csvfile:
      writer = csv.writer(csvfile, delimiter='\t', quotechar='', quoting=csv.QUOTE_NONE)
      writer.writerow(['id','whg_pid','title','ccodes','lon','lat','added','matches'])
      for f in features:
        geoms = f.geoms.all()
        gobj = augGeom(geoms)
        row = [
          str(f.src_id),
          str(f.id),f.title,
          ';'.join(f.ccodes),
          gobj['lonlat'][0] if 'lonlat' in gobj else None,
          gobj['lonlat'][1] if 'lonlat' in gobj else None,
          gobj['new'] if 'new' in gobj else None,
          str(augLinks(f.links.all())) ]
        writer.writerow(row)
        #print(row)
    response = FileResponse(open(fn, 'rb'),content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="'+os.path.basename(fn)+'"'

    return response
  else:
    # make file name
    fn = 'media/user_'+user+'/'+ds.label+'_aug_'+date+'.json'

    with open(fn, 'w', encoding='utf-8') as outfile:
      #fcoll = {"type":"FeatureCollection","features":[]}
      for f in features:
        print('dl_aug, lpf adding feature:',f)
        feat={"type":"Feature",
              "properties":{
                "@id":f.dataset.uri_base+f.src_id,
                "src_id":f.src_id,
                "title":f.title,
                "whg_pid":f.id}}
        if len(f.geoms.all()) >1:
          feat['geometry'] = {'type':'GeometryCollection'}
          feat['geometry']['geometries'] = [g.jsonb for g in f.geoms.all()]
        elif len(f.geoms.all()) == 1:
          feat['geometry'] = f.geoms.first().jsonb
        else: # no geoms
          feat['geometry'] = feat['geometry'] = {'type':'GeometryCollection','geometries':[]}
        feat['names'] = [n.jsonb for n in f.names.all()]
        feat['types'] = [t.jsonb for t in f.types.all()]
        feat['when'] = [w.jsonb for w in f.whens.all()]
        feat['relations'] = [r.jsonb for r in f.related.all()]
        feat['links'] = [l.jsonb for l in f.links.all()]
        feat['descriptions'] = [des.jsonb for des in f.descriptions.all()]
        feat['depictions'] = [dep.jsonb for dep in f.depictions.all()]
        #fcoll['features'].append(feat)  
        outfile.write(json.dumps(feat,indent=2))
      #outfile.write(json.dumps(fcoll,indent=2))

    # response is reopened file
    response = FileResponse(open(fn, 'rb'), content_type='text/json')
    response['Content-Disposition'] = 'attachment; filename="'+os.path.basename(fn)+'"'

    return response

# ***
# format tsv validation errors for modal display
# ***
def parse_errors_tsv(err):
  #print('parse_errors_tsv()',err)
  html = '<p>Nah, errors in that file...'
  return html

# *** NOT YET CALLED
# format lpf validation errors for modal display
# *** 
def parse_errors_lpf(err):
  print('parse_errors_lpf()',err)

# 
# validate Linked Places json-ld (w/jsonschema)
# format ['coll' (FeatureCollection) | 'lines' (json-lines)]
def validate_lpf(tempfn,format):
  # TODO: handle json-lines
  # TODO: create v1.1 schema; phase out v1.0
  #wd = '/Users/karlg/Documents/Repos/_whgazetteer/'
  schema = json.loads(codecs.open('datasets/static/validate/schema_lpf_v1.1.json','r','utf8').read())
  # rename tempfn
  newfn = tempfn+'.jsonld'
  os.rename(tempfn,newfn)
  #infile = codecs.open(tempfn, 'r', 'utf8')
  infile = codecs.open(newfn, 'r', 'utf8')
  #fout = codecs.open('validate-lpf-result.txt','w','utf8')
  result = {"format":"lpf","errors":[]}
  [countrows,count_ok] = [0,0]

  jdata = json.loads(infile.read())
  if ['type', '@context', 'features'] != list(jdata.keys()) \
     or jdata['type'] != 'FeatureCollection' \
     or len(jdata['features']) == 0:
    print('not valid GeoJSON-LD')
  else:
    for feat in jdata['features']:
      countrows +=1
      #print(feat['properties']['title'])
      try:
        validate(
          instance=feat,
          schema=schema,
          format_checker=draft7_format_checker
        )
        count_ok +=1
      except:
        err = sys.exc_info()
        print('some kinda error',err)
        result["errors"].append({"feat":countrows,'error':err[1].args[0]})

  #fout.write(json.dumps(result["errors"]))
  #fout.close()
  result['count'] = countrows
  #print('validate_lpf() result',result)
  return result

#
# validate LP-TSV file (w/goodtables library)
def validate_tsv(tempfn):
  result = {"errors":[],"format":"delimited"}
  # TODO: detect encoding
  newfn = tempfn+'.tsv'
  os.rename(tempfn,newfn)
  print('tempfn,newfn',tempfn,type(tempfn),newfn,type(newfn))
  schema_lptsv = json.loads(codecs.open('datasets/static/validate/schema_tsv.json', 'r', 'utf8').read())
  report = gvalidate(newfn,schema=schema_lptsv,order_fields=True,row_limit=20000,skip_checks='blank-header')
  pp.pprint(report)  
  #print('error count',report['error-count'])
  result['count'] = report['tables'][0]['row-count']-1 # counts header apparently
  result['columns'] = report['tables'][0]['headers']
  result['file'] = report['tables'][0]['source']
  # make sense of errors for users
  # filter harmless errors (a goodtable library bug IMO)
  errors = [x for x in report['tables'][0]['errors'] if x['code'] not in ["blank-header","missing-header"]]
  error_types = list(set([x['code'] for x in errors]))
  if 'non-matching-header' in error_types:
    # have user fix that issue and try again
    #result['errors'].append({'message':'One or more column heading is invalid: '+str(result['columns'])})
    result['errors'].append('One or more column heading is invalid: '+str(result['columns']))
  else:
    result['errors'] = [x['message'].replace('and format "default"','') for x in errors]
  return result

class HitRecord(object):
  #def __init__(self, whg_id, place_id, dataset, src_id, title):
  def __init__(self, place_id, dataset, src_id, title):
    #self.whg_id = whg_id
    self.place_id = place_id
    self.src_id = src_id
    self.title = title
    self.dataset = dataset

  def __str__(self):
    import json
    return json.dumps(str(self.__dict__))    
    #return json.dumps(self.__dict__)

  def toJSON(self):
    import json
    return json.loads(json.dumps(self.__dict__,indent=2))


def aat_lookup(id):
  try:
    label = aat.types[id]['term_full']
    return label
  except:
    print(id,' broke it')
    print("error:", sys.exc_info())        

def hully(g_list):
  from django.contrib.gis.geos import GEOSGeometry
  from django.contrib.gis.geos import MultiPoint
  from django.contrib.gis.geos import GeometryCollection

  # maybe mixed bag
  types = list(set([g['type'] for g in g_list]))
  # should work for any combination
  try:
    hull=GeometryCollection([GEOSGeometry(json.dumps(g)) for g in g_list]).convex_hull
  except:
    print('hully() failed on g_list',g_list)
  # buffer points, but only a little if near meridian
  if len(types) ==1 and types[0] == 'Point':
    l = list(set([g_list[0]['coordinates'][0] for c in g_list[0]]))
    if len([i for i in l if i >= 175]) == 0:
      hull = hull.buffer(1)
    else:
      hull = hull.buffer(0.1)
  return json.loads(hull.geojson) if hull.geojson !=None else []

def parse_wkt(g):
  #print('wkt',g)
  from shapely.geometry import mapping
  gw = wkt.loads(g)
  feature = mapping(gw)
  #print('wkt, feature',g, feature)
  return feature

#
def maketime():
  ts = time.time()
  sttime = datetime.datetime.fromtimestamp(ts).strftime('%Y%m%d_%H%M%S')
  return sttime
def myprojects(me):
  return DatasetUser.objects.filter(user_id_id=me.id).values_list('dataset_id_id')
def parsejson(value,key):
  """returns value for given key"""
  print('parsejson() value',value)
  obj = json.loads(value.replace("'",'"'))
  return obj[key]
def makeCoords(lonstr,latstr):
  #print(type(lonstr),latstr)
  lon = float(lonstr) if lonstr not in ['','nan'] else ''
  lat = float(latstr) if latstr not in ['','nan'] else ''
  #lon = float(lonstr) if lonstr != '' else ''
  #lat = float(latstr) if latstr != '' else ''
  coords = [] if (lonstr == ''  or latstr == '') else [lon,lat]
  return coords
def elapsed(delta):
  minutes, seconds = divmod(delta.seconds, 60)
  return '{:02}:{:02}'.format(int(minutes), int(seconds))

def bestParent(qobj, flag=False):
  # applicable for tgn only 
  best = []
  print('qobj in bestParent',qobj)
  # merge parent country/ies & parents
  if len(qobj['countries']) > 0 and qobj['countries'][0] != '':
    for c in qobj['countries']:
      best.append(parents.ccodes[0][c]['tgnlabel'])
  if len(qobj['parents']) > 0:
    for p in qobj['parents']:
      best.append(p)
  if len(best) == 0:
    best = ['World']
  return best

# wikidata Qs from ccodes
#TODO: consolidate hashes
def getQ(arr,what):
  print('arr,what',arr, what)
  qids=[]
  if what == 'ccodes':
    from datasets.static.hashes.parents import ccodes
    for c in arr:
      qids.append('wd:'+ccodes[0][c]['wdid'])
  elif what == 'types':
    if len(arr) == 0:
      qids.append('wd:Q486972')
    for t in arr:
      if t in aat_q.qnums:
        for q in aat_q.qnums[t]:
          qids.append('wd:'+q)
      else:
        qids.append('wd:Q486972')
  return qids

def roundy(x, direct="up", base=10):
  import math
  if direct == "down":
    return int(math.ceil(x / 10.0)) * 10 - base
  else:
    return int(math.ceil(x / 10.0)) * 10

def fixName(toponym):
  import re
  search_name = toponym
  r1 = re.compile(r"(.*?), Gulf of")
  r2 = re.compile(r"(.*?), Sea of")
  r3 = re.compile(r"(.*?), Cape")
  r4 = re.compile(r"^'")
  if bool(re.search(r1,toponym)):
    search_name = "Gulf of " + re.search(r1,toponym).group(1)
  if bool(re.search(r2,toponym)):
    search_name = "Sea of " + re.search(r2,toponym).group(1)
  if bool(re.search(r3,toponym)):
    search_name = "Cape " + re.search(r3,toponym).group(1)
  if bool(re.search(r4,toponym)):
    search_name = toponym[1:]
  return search_name if search_name != toponym else toponym

# in: list of Black atlas place types
# returns list of equivalent classes or types for {gaz}
def classy(gaz, typeArray):
  import codecs, json
  #print(typeArray)
  types = []
  finhash = codecs.open('../data/feature-classes.json', 'r', 'utf8')
  classes = json.loads(finhash.read())
  finhash.close()
  if gaz == 'gn':
    t = classes['geonames']
    default = 'P'
    for k,v in t.items():
      if not set(typeArray).isdisjoint(t[k]):
        types.append(k)
      else:
        types.append(default)
  elif gaz == 'tgn':
    t = classes['tgn']
    default = 'inhabited places' # inhabited places
    # if 'settlement' exclude others
    typeArray = ['settlement'] if 'settlement' in typeArray else typeArray
    # if 'admin1' (US states) exclude others
    typeArray = ['admin1'] if 'admin1' in typeArray else typeArray
    for k,v in t.items():
      if not set(typeArray).isdisjoint(t[k]):
        types.append(k)
      else:
        types.append(default)
  elif gaz == "dbp":
    t = classes['dbpedia']
    default = 'Place'
    for k,v in t.items():
      # is any Black type in dbp array?
      # TOD: this is crap logic, fix it
      if not set(typeArray).isdisjoint(t[k]):
        types.append(k)
      #else:
        #types.append(default)
  if len(types) == 0:
    types.append(default)
  return list(set(types))

class UpdateCountsView(View):
  """ Returns counts of unreviewed hits per pass """
  @staticmethod
  def get(request):
    print('UpdateCountsView GET:',request.GET)
    """
    args in request.GET:
        [integer] ds_id: dataset id
    """
    ds = get_object_or_404(Dataset, id=request.GET.get('ds_id'))
    updates = {}
    for tid in [t.task_id for t in ds.tasks.all()]:
      updates[tid] = {
        'pass1':len(Hit.objects.raw('select distinct on (place_id_id) place_id_id, id from hits where task_id = %s and query_pass=%s and reviewed=false',[tid,'pass1'])),
        'pass2':len(Hit.objects.raw('select distinct on (place_id_id) place_id_id, id from hits where task_id = %s and query_pass=%s and reviewed=false',[tid,'pass2'])),
        'pass3':len(Hit.objects.raw('select distinct on (place_id_id) place_id_id, id from hits where task_id = %s and query_pass=%s and reviewed=false',[tid,'pass3']))
      }    
    return JsonResponse(updates, safe=False)

# ABANDONED for views.ds_compare
#def compare(dsid):
  ## get last two uploaded DatasetFile instances for this dataset
  ## a is the current data, b is newly uploaded
  ##file_a = DatasetFile.objects.filter(dataset_id_id=dsid).order_by('-rev')[1]
  ## TODO: for testing, a is always the initial
  #file_a = get_object_or_404(DatasetFile,dataset_id_id=dsid,rev=1)
  #file_b = DatasetFile.objects.filter(dataset_id_id=dsid).order_by('-rev')[0]
  #fn_a = file_a.file.name
  #fn_b = file_b.file.name
  #print('dsid; files to compare():',dsid,fn_a,fn_b)
  #print('files exist?',Path('media/'+fn_a).exists(),Path('media/'+fn_b).exists())
  #adf = pd.read_csv('media/'+fn_a,delimiter='\t')
  ##print('adf',adf)
  #bdf = pd.read_csv('/Users/karlg/Documents/Repos/_whgazetteer/media/'+fn_b,delimiter='\t')
  ##try:
    ##bdf = pd.read_csv('media/'+fn_b,delimiter='\t')
  ##except:
    ##sys.exit(sys.exc_info())
  #ids_a = adf['id'].tolist()
  #ids_b = bdf['id'].tolist()
  #resobj={"count_new":len(ids_b),'count_diff':len(ids_b)-len(ids_a)}
  ## new or removed columns?
  #col_del = list(set(adf.columns)-set(bdf.columns))
  #col_add = list(set(bdf.columns)-set(adf.columns))
  #resobj['col_add']=col_add
  #resobj['col_del']=col_del
  #resobj['rows_add']=list(set(ids_b)-set(ids_a))
  #resobj['rows_del']=list(set(ids_a)-set(ids_b))
  ## build a little blurb
  #text='The revised dataset has '+str(resobj['count_new'])+' records, a difference of '+str(resobj['count_diff'])+". Columns "
  #text += 'to add: '+str(resobj['col_add']) + '. 'if len(resobj['col_add']) > 0 else \
    #'to remove: '+ str(resobj['col_del'])+'. ' if len(resobj['col_del']) > 0 \
          #else "remain the same. "
  #text += 'Records to be added: '+str(resobj['rows_add'])+'. ' if len(resobj['rows_add'])>0 else ''
  #text += 'Records to be removed: '+str(resobj['rows_del'])+'. ' if len(resobj['rows_del'])>0 else ''
  #text += 'All records with an ID matching one in the existing dataset will be replaced.'  
  #resobj['text'] = text

  #return resobj
